# app/services/excel_generator_service.py

import io
from datetime import datetime
from typing import List, Dict, Any, Optional, Set # Añadido Set

# Importamos la librería de Excel
import openpyxl 
from openpyxl.styles import Alignment, Border, Side, Font 

# FastAPI / SQLModel
from sqlmodel import Session, select, func
from fastapi import HTTPException, status

# Models
from app.models.tutor import Tutor
from app.models.tutoria import Tutoria, EstadoTutoria
from app.models.alumno import Alumno
from app.models.reporte_integral import ReporteIntegral

# Ruta a la plantilla
TEMPLATE_ANEXO3_PATH = "app/excel_templates/ANEXO_3_formato.xlsx" 

def generate_anexo3_reporte(db: Session, periodo: str) -> io.BytesIO:
    """
    Genera el reporte "ANEXO 3" rellenando Hoja 2 (Detalle) y Hoja 1 (Resumen)
    con datos de la BD de forma dinámica.
    """
    
    # --- 1. Obtener Datos de la BD ---
    query = (
        select(Tutoria, Alumno, ReporteIntegral)
        .join(Alumno, Tutoria.alumno_id == Alumno.id_alumno) # type: ignore
        .outerjoin(ReporteIntegral, Tutoria.id_tutoria == ReporteIntegral.id_tutoria) # type: ignore
        .where(Tutoria.periodo == periodo)
        .order_by(Alumno.apellido_p, Alumno.apellido_m, Alumno.nombre) # type: ignore
    )
    resultados = db.exec(query).all()

    if not resultados:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No se encontraron tutorías asignadas para el periodo {periodo}."
        )

    # --- 2. Cargar la Plantilla de Excel ---
    try:
        workbook = openpyxl.load_workbook(TEMPLATE_ANEXO3_PATH)
    except FileNotFoundError:
        raise HTTPException(status_code=500, detail=f"No se encontró la plantilla Excel en: {TEMPLATE_ANEXO3_PATH}")

    # --- 3. Definir Estilos Comunes ---
    thin_side = Side(style="thin", color="000000")
    all_borders = Border(left=thin_side, top=thin_side, right=thin_side, bottom=thin_side)
    bold_font = Font(bold=True)
    
    # --- 4. Estructura para acumular datos del Resumen (Hoja 1) ---
    resumen_data: Dict[tuple, Dict[str, Any]] = {}
    
    # --- 5. Rellenar Hoja 2 (Detalle) y Acumular para Hoja 1 ---
    try:
        sheet2 = workbook["Hoja2"]
    except KeyError:
        raise HTTPException(status_code=500, detail="La plantilla no contiene una hoja llamada 'Hoja2'.")

    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    sheet2["K14"] = fecha_actual

    FILA_INICIO_H2 = 18
    fila_actual_h2 = FILA_INICIO_H2

    for index, (tutoria, alumno, reporte) in enumerate(resultados):
        
        # --- A. Preparar variables auxiliares (para Hoja 1 y Hoja 2) ---
        r_grupal = 0
        r_individual = 0
        r_psicologia = 0
        r_ciencias = 0
        r_jefatura = 0

        if reporte:
            r_grupal = reporte.tutoria_grupal
            r_individual = reporte.tutoria_individual
            r_psicologia = 1 if reporte.psicologia > 0 else 0
            r_ciencias = 1 if reporte.ciencias_basicas > 0 else 0
            r_jefatura = 1 if reporte.jefatura_academica > 0 else 0

        # --- B. Llenar Hoja 2 (Detalle) ---
        sheet2[f'A{fila_actual_h2}'] = f"{alumno.apellido_p} {alumno.apellido_m or ''} {alumno.nombre}".strip()
        sheet2[f'B{fila_actual_h2}'] = alumno.num_control
        sheet2[f'C{fila_actual_h2}'] = r_grupal
        sheet2[f'D{fila_actual_h2}'] = r_individual
        sheet2[f'J{fila_actual_h2}'] = r_psicologia + r_ciencias + r_jefatura # Suma de canalizaciones por alumno
        sheet2[f'K{fila_actual_h2}'] = r_psicologia
        sheet2[f'Q{fila_actual_h2}'] = r_ciencias
        sheet2[f'R{fila_actual_h2}'] = r_jefatura

        # Aplicar bordes Hoja 2
        for col_num in range(1, 19): # Col A a R
            cell = sheet2.cell(row=fila_actual_h2, column=col_num)
            cell.border = all_borders
        
        fila_actual_h2 += 1

        # --- C. Acumular Datos para Hoja 1 (Resumen) ---
        clave = (alumno.carrera, alumno.semestre_actual)
        
        if clave not in resumen_data:
            resumen_data[clave] = {
                "tutores_ids": set(),
                "total_grupal": 0,
                "total_individual": 0,
                "total_canalizaciones": 0,
                "total_psicologia": 0,    # <-- NUEVO ACUMULADOR
                "total_ciencias": 0,      # <-- NUEVO ACUMULADOR
                "total_jefatura": 0       # <-- NUEVO ACUMULADOR
            }
        
        # Actualizar acumuladores
        resumen_data[clave]["tutores_ids"].add(tutoria.tutor_id)
        resumen_data[clave]["total_grupal"] += r_grupal
        resumen_data[clave]["total_individual"] += r_individual
        resumen_data[clave]["total_canalizaciones"] += (r_psicologia + r_ciencias + r_jefatura)
        resumen_data[clave]["total_psicologia"] += r_psicologia  # <-- NUEVA LÓGICA
        resumen_data[clave]["total_ciencias"] += r_ciencias      # <-- NUEVA LÓGICA
        resumen_data[clave]["total_jefatura"] += r_jefatura      # <-- NUEVA LÓGICA

    # --- D. Totales Hoja 2 (Fórmulas) ---
    fila_total_h2 = fila_actual_h2
    fila_fin_datos_h2 = fila_actual_h2 - 1
    sheet2[f'A{fila_total_h2}'] = "TOTAL"
    sheet2.merge_cells(f'A{fila_total_h2}:B{fila_total_h2}')
    sheet2[f'C{fila_total_h2}'] = f"=SUM(C{FILA_INICIO_H2}:C{fila_fin_datos_h2})"
    sheet2[f'D{fila_total_h2}'] = f"=SUM(D{FILA_INICIO_H2}:D{fila_fin_datos_h2})"
    sheet2[f'J{fila_total_h2}'] = f"=SUM(J{FILA_INICIO_H2}:J{fila_fin_datos_h2})"
    sheet2[f'K{fila_total_h2}'] = f"=SUM(K{FILA_INICIO_H2}:K{fila_fin_datos_h2})" # Contar 1s es lo mismo que sumar
    sheet2[f'Q{fila_total_h2}'] = f"=SUM(Q{FILA_INICIO_H2}:Q{fila_fin_datos_h2})" # Contar 1s es lo mismo que sumar
    sheet2[f'R{fila_total_h2}'] = f"=SUM(R{FILA_INICIO_H2}:R{fila_fin_datos_h2})" # Contar 1s es lo mismo que sumar

    for col_num in range(1, 19):
        cell = sheet2.cell(row=fila_total_h2, column=col_num)
        cell.border = all_borders
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ==========================================
    # === PROCESAMIENTO HOJA 1 (RESUMEN) ===
    # ==========================================
    try:
        sheet1 = workbook["Hoja1"]
    except KeyError:
        raise HTTPException(status_code=500, detail="La plantilla no contiene una hoja llamada 'Hoja1'.")

    # --- NUEVO: Rellenar Fecha en Hoja 1 ---
    sheet1["L4"] = fecha_actual # Celda K14 en Hoja2, L4 en Hoja1

    FILA_INICIO_H1 = 18
    fila_actual_h1 = FILA_INICIO_H1
    
    claves_ordenadas = sorted(resumen_data.keys(), key=lambda x: (x[0], x[1]))

    for carrera, semestre in claves_ordenadas:
        datos = resumen_data[(carrera, semestre)]
        
        # --- Llenar Celdas (CON NUEVOS CAMPOS) ---
        sheet1[f'A{fila_actual_h1}'] = carrera
        sheet1[f'B{fila_actual_h1}'] = semestre
        sheet1[f'C{fila_actual_h1}'] = len(datos["tutores_ids"])
        sheet1[f'E{fila_actual_h1}'] = datos["total_grupal"]
        sheet1[f'F{fila_actual_h1}'] = datos["total_individual"]
        sheet1[f'R{fila_actual_h1}'] = datos["total_canalizaciones"]
        sheet1[f'S{fila_actual_h1}'] = datos["total_psicologia"]    # <-- NUEVA CELDA
        sheet1[f'Y{fila_actual_h1}'] = datos["total_ciencias"]      # <-- NUEVA CELDA
        sheet1[f'Z{fila_actual_h1}'] = datos["total_jefatura"]      # <-- NUEVA CELDA

        # Aplicar bordes (extendido hasta Col Z = 26)
        # 28 -> 27 (AA) -> 26 (Z)
        # Range(1, 27) -> Col 1 a 26
        # Range(1, 30) -> Col 1 a 29 (AC) (para cubrir todo el formato visible)
        for col_num in range(1, 28): 
            cell = sheet1.cell(row=fila_actual_h1, column=col_num)
            cell.border = all_borders
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_num == 1: # Alinear carrera a la izquierda
                cell.alignment = Alignment(horizontal="left", vertical="center")

        fila_actual_h1 += 1

    # --- Totales Hoja 1 (Fórmulas CON NUEVOS CAMPOS) ---
    fila_total_h1 = fila_actual_h1
    fila_fin_datos_h1 = fila_actual_h1 - 1
    
    sheet1[f'A{fila_total_h1}'] = "TOTAL"
    sheet1.merge_cells(f'A{fila_total_h1}:B{fila_total_h1}') 
    
    sheet1[f'C{fila_total_h1}'] = f"=SUM(C{FILA_INICIO_H1}:C{fila_fin_datos_h1})"
    sheet1[f'E{fila_total_h1}'] = f"=SUM(E{FILA_INICIO_H1}:E{fila_fin_datos_h1})"
    sheet1[f'F{fila_total_h1}'] = f"=SUM(F{FILA_INICIO_H1}:F{fila_fin_datos_h1})"
    sheet1[f'R{fila_total_h1}'] = f"=SUM(R{FILA_INICIO_H1}:R{fila_fin_datos_h1})"
    sheet1[f'S{fila_total_h1}'] = f"=SUM(S{FILA_INICIO_H1}:S{fila_fin_datos_h1})" # <-- NUEVA FÓRMULA
    sheet1[f'Y{fila_total_h1}'] = f"=SUM(Y{FILA_INICIO_H1}:Y{fila_fin_datos_h1})" # <-- NUEVA FÓRMULA
    sheet1[f'Z{fila_total_h1}'] = f"=SUM(Z{FILA_INICIO_H1}:Z{fila_fin_datos_h1})" # <-- NUEVA FÓRMULA

    # Estilos Fila Total Hoja 1 (extendido hasta Col Z)
    for col_num in range(1, 28):
        cell = sheet1.cell(row=fila_total_h1, column=col_num)
        cell.border = all_borders
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- 8. Guardar el Excel en Memoria ---
    output_stream = io.BytesIO()
    workbook.save(output_stream)
    workbook.close()
    
    output_stream.seek(0)
    return output_stream

# --- FIN DEL ARCHIVO ---