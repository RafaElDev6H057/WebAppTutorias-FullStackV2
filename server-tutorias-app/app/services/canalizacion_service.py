import io
from typing import List, Tuple

# Importamos librería para crear Excel desde cero
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# FastAPI / SQLModel
from sqlmodel import Session, select
from fastapi import HTTPException, status

# Modelos necesarios
from app.models.alumno import Alumno
from app.models.tutoria import Tutoria
from app.models.reporte_integral import ReporteIntegral

# --- FUNCIÓN PRIVADA AUXILIAR PARA GENERAR EL EXCEL (DRY Principle) ---
def _generar_excel_comun(resultados: List[Tuple[Alumno, Tutoria, ReporteIntegral]], titulo_reporte: str) -> io.BytesIO:
    """
    Función interna que recibe los resultados de la BD y genera el archivo Excel
    con el formato estandarizado para canalizaciones.
    """
    # Crear el libro de Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Canalización" # type: ignore

    # --- Estilos ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Azul institucional
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_side = Side(style="thin", color="000000")
    all_borders = Border(left=thin_side, top=thin_side, right=thin_side, bottom=thin_side)

    # --- Encabezados ---
    headers = ["Nombre Completo", "No. Control", "Carrera", "Semestre", "Correo", "Teléfono"]
    
    # Título superior
    sheet.merge_cells('A1:F1') # type: ignore
    title_cell = sheet['A1'] # type: ignore
    title_cell.value = titulo_reporte
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_alignment

    # Fila de Encabezados (Fila 3)
    row_header = 3
    for col_idx, header_text in enumerate(headers, start=1):
        cell = sheet.cell(row=row_header, column=col_idx, value=header_text) # type: ignore
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = all_borders

    # --- Rellenar Datos ---
    row_idx = 4
    for alumno, tutoria, reporte in resultados:
        # Construir nombre completo
        nombre_completo = f"{alumno.apellido_p} {alumno.apellido_m or ''} {alumno.nombre}".strip()
        
        datos_alumno = [
            nombre_completo,
            alumno.num_control,
            alumno.carrera,
            alumno.semestre_actual,
            alumno.correo or "Sin correo",
            alumno.telefono or "Sin teléfono"
        ]

        for col_idx, valor in enumerate(datos_alumno, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=valor) # type: ignore
            cell.border = all_borders
            
            # Ajuste de alineación: Control, Carrera, Semestre centrados
            if col_idx in [2, 3, 4]: 
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
        
        row_idx += 1

    # --- Ajustar ancho de columnas ---
    column_widths = [40, 15, 42, 10, 35, 35] 
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        sheet.column_dimensions[col_letter].width = width # type: ignore

    # --- Guardar ---
    output_stream = io.BytesIO()
    workbook.save(output_stream)
    workbook.close()
    
    output_stream.seek(0)
    return output_stream


# --- SERVICIOS PÚBLICOS ---

def generate_reporte_psicologia(db: Session, periodo: str) -> io.BytesIO:
    """Genera reporte para PSICOLOGÍA"""
    query = (
        select(Alumno, Tutoria, ReporteIntegral)
        .join(Tutoria, Alumno.id_alumno == Tutoria.alumno_id) # type: ignore
        .join(ReporteIntegral, Tutoria.id_tutoria == ReporteIntegral.id_tutoria) # type: ignore
        .where(Tutoria.periodo == periodo)
        .where(ReporteIntegral.psicologia > 0)
        .order_by(Alumno.carrera, Alumno.semestre_actual, Alumno.apellido_p) # type: ignore
    )
    resultados = db.exec(query).all()

    if not resultados:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No se encontraron alumnos canalizados a Psicología en el periodo {periodo}."
        )

    titulo = f"REPORTE DE CANALIZACIÓN - PSICOLOGÍA - PERIODO {periodo}"
    return _generar_excel_comun(resultados, titulo) #type: ignore


def generate_reporte_ciencias_basicas(db: Session, periodo: str) -> io.BytesIO:
    """Genera reporte para CIENCIAS BÁSICAS"""
    query = (
        select(Alumno, Tutoria, ReporteIntegral)
        .join(Tutoria, Alumno.id_alumno == Tutoria.alumno_id) # type: ignore
        .join(ReporteIntegral, Tutoria.id_tutoria == ReporteIntegral.id_tutoria) # type: ignore
        .where(Tutoria.periodo == periodo)
        .where(ReporteIntegral.ciencias_basicas > 0) # Filtro específico
        .order_by(Alumno.carrera, Alumno.semestre_actual, Alumno.apellido_p) # type: ignore
    )
    resultados = db.exec(query).all()

    if not resultados:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No se encontraron alumnos canalizados a Ciencias Básicas en el periodo {periodo}."
        )

    titulo = f"REPORTE DE CANALIZACIÓN - CIENCIAS BÁSICAS - PERIODO {periodo}"
    return _generar_excel_comun(resultados, titulo) #type: ignore


def generate_reporte_jefatura_academica(db: Session, periodo: str) -> io.BytesIO:
    """Genera reporte para JEFATURA ACADÉMICA"""
    query = (
        select(Alumno, Tutoria, ReporteIntegral)
        .join(Tutoria, Alumno.id_alumno == Tutoria.alumno_id) # type: ignore
        .join(ReporteIntegral, Tutoria.id_tutoria == ReporteIntegral.id_tutoria) # type: ignore
        .where(Tutoria.periodo == periodo)
        .where(ReporteIntegral.jefatura_academica > 0) # Filtro específico
        .order_by(Alumno.carrera, Alumno.semestre_actual, Alumno.apellido_p) # type: ignore
    )
    resultados = db.exec(query).all()

    if not resultados:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No se encontraron alumnos canalizados a Jefatura Académica en el periodo {periodo}."
        )

    titulo = f"REPORTE DE CANALIZACIÓN - JEFATURA ACADÉMICA - PERIODO {periodo}"
    return _generar_excel_comun(resultados, titulo) #type: ignore